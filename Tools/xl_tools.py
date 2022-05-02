import openpyxl
from pathlib import Path
import sentence_analysis
from AI_Assistant import Aiyu
from Language_Data import file_names


def read_from_xl(num_lines):
    xlsx_file = Path("/Users/mushr/PycharmProjects/AI_Law_Assistant/Language_Data/Training_raw/问题筛选和分类1.xlsx")
    wb_file = openpyxl.load_workbook(xlsx_file)
    # print(wb_file.sheetnames)  # print all sheet names
    # Create the bot for putting the sentence into the universal format.
    universal_bot = Aiyu.Aiyu(file_names.jsn_cbot_ch_n, 'intents', 'category', 'patterns', 'responses', model_name='ORG_FORM')
    universal_bot.data_processor("AIYU_PKL", force_process="N", split_mode='Y')
    universal_bot.construct_softmax_model(16, 8, 120, "AIYU_Core", retrain_model='N')

    data_sheet = wb_file['sheet0']
    for row in data_sheet.iter_rows(max_row=num_lines):
        for cell in row:
            if cell.value:
                print(cell.value)
            else:
                continue




read_from_xl(2)